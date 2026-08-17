from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from datetime import datetime
from io import BytesIO
from typing import List, Optional

import openpyxl
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.limiter import limiter
from app.features.auth.dependencies import get_current_user, require_admin
from app.features.leaves.schemas.leaves import (
    LeaveApplyRequest, 
    LeaveHistoryItem, 
    EmployeeLeaveBalanceSummary,
    LeaveTypeOption,
    Holiday,
    BatchLeaveRequest,
    CancelLeaveRequest,
    CancelLeaveRejectRequest,
    BatchCancelLeaveRejectRequest,
)
from app.features.leaves.services.leave_service import LeaveBusinessService
from app.middleware.logging import _get_client_ip
from app.models.employee_master import EmployeeMaster
from app.models.leave import Leave
from app.models.leave_type import LeaveTypeConfig
from app.models.holiday import Holiday as HolidayModel
from app.services.database.dependencies import get_leave_business_service
from app.features.redmine.service import redmine_service
from app.utils.audit import audit_logger
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/leaves", tags=["leaves"])

@router.get("/balance", response_model=EmployeeLeaveBalanceSummary)
async def get_balance(
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    return await leave_service.get_employee_leave_balance_summary(current_user.get("sub"))

@router.get("/leave-types", response_model=List[LeaveTypeOption])
@limiter.limit("30/minute")
async def list_active_leave_types(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    types = (
        db.query(LeaveTypeConfig)
        .filter(LeaveTypeConfig.is_active.is_(True))
        .order_by(LeaveTypeConfig.name)
        .all()
    )
    return [
        LeaveTypeOption(id=str(t.id), code=t.code, name=t.name, is_paid=t.is_paid)
        for t in types
    ]

@router.get("/self-leave-balance")
@limiter.limit("30/minute")
async def get_self_leave_balance(
    request: Request,
    fiscal_year: int = Query(None, ge=2000, description="Fiscal year (default current)"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    return await leave_service.get_self_leave_balance(current_user, fiscal_year)

@router.get("/self-leave-balance-grid")
@limiter.limit("30/minute")
async def get_self_leave_balance_grid(
    request: Request,
    fiscal_year: int = Query(None, ge=2000, description="Fiscal year (default current)"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    return await leave_service.get_self_leave_balance_grid(current_user, fiscal_year)

@router.get("/admin/{email}", response_model=List[LeaveHistoryItem])
async def get_user_leave_history(
    email: str,
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can view leave history."
        )
    return await leave_service.get_user_leave_history(email, current_user, from_date, to_date)

@router.get("/users")
async def list_leave_users(
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    """Get list of users whose leaves can be viewed.
    - Admin: All Redmine users.
    - PM: Users sharing at least one project.
    """
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can view team leave history."
        )
    return await leave_service.get_visible_leave_users(current_user)

@router.get("/history")
async def get_history(
    team: Optional[bool] = Query(None, description="PM/PC: view all team members' leaves"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    user_id = current_user.get("sub")
    email = current_user.get("email")
    roles = current_user.get("roles", [])

    if team:
        if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
            raise HTTPException(status_code=403, detail="Admin, PM, or PC access required.")
        from app.features.redmine.sql_service import RedmineSQLService
        sql = RedmineSQLService(leave_service.leave_db.db)
        rm_user = sql.get_user_by_email(email)
        if not rm_user:
            raise HTTPException(status_code=404, detail="Current user not found in Redmine.")
        member_rm_ids = sql.get_team_member_ids(rm_user["id"])
        member_rm_ids.add(rm_user["id"])
        emp_q = leave_service.leave_db.db.query(EmployeeMaster).filter(
            EmployeeMaster.redmine_user_id.in_(member_rm_ids)
        ).all()
        member_emails = [e.user_email for e in emp_q if e.user_email]
        if not member_emails:
            return {"total": 0, "records": []}
        records = await leave_service.get_team_leaves(member_emails, from_date, to_date)
        return {"total": len(records), "records": records}

    records = await leave_service.get_leave_history(user_id, from_date, to_date)
    return {"total": len(records), "records": records}


def _build_leaves_excel(records: list) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leaves"
    ws.append(["Date", "Day", "Employee", "Designation", "Leave Type", "Status",
               "Days", "Reason", "Contact", "Is Traveling"])

    def _fmt(v):
        if v is None:
            return ""
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return str(v)

    for r in records:
        days = ""
        if r.get("start_date") and r.get("end_date"):
            try:
                days = (r["end_date"] - r["start_date"]).days + 1
            except TypeError:
                days = ""
        ws.append([
            _fmt(r.get("start_date")),
            r["start_date"].strftime("%A") if r.get("start_date") and hasattr(r["start_date"], "strftime") else "",
            r.get("userName") or "",
            r.get("userDesignation") or "",
            r.get("leave_type") or "",
            r.get("status") or "",
            days,
            r.get("reason") or "",
            r.get("contact_number") or "",
            "Yes" if r.get("is_traveling") else "No",
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/my-leaves")
async def get_my_leaves(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    sort_by: str = Query("start_date", description="start_date, end_date, leave_type, status, userName"),
    sort_dir: str = Query("asc", description="asc or desc"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    leave_type: Optional[str] = Query(None, description="Comma-separated: EL,PL,UPL"),
    status: Optional[str] = Query(None, description="Comma-separated: pending,approved,rejected,emergency,cancelled,cancellation_requested,cancellation_rejected"),
    search: Optional[str] = Query(None, description="Search reason, comment, leave type, status"),
    export: bool = Query(False, description="Export as Excel spreadsheet"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service),
):
    email = current_user.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="User email not found.")
    result = await leave_service.get_leaves_paginated(
        emails=[email],
        from_date=from_date, to_date=to_date,
        leave_type=leave_type, status=status,
        search=search, include_name_search=False,
        sort_by=sort_by, sort_dir=sort_dir,
        page=page, page_size=page_size, export=export,
    )
    if export:
        output = _build_leaves_excel(result["records"])
        audit_logger.info(
            f"Leave history exported by {current_user.get('username')}",
            extra={
                "correlation_id": request.state.correlation_id,
                "extra_data": {
                    "action": "export_my_leaves",
                    "username": current_user.get("username"),
                    "status": "success",
                    "client_ip": _get_client_ip(request),
                    "exported_rows": len(result["records"]),
                },
            },
        )
        filename = f"my_leaves_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    return result


@router.get("/team-leaves")
async def get_team_leaves(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    sort_by: str = Query("start_date", description="start_date, end_date, leave_type, status, userName"),
    sort_dir: str = Query("asc", description="asc or desc"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    leave_type: Optional[str] = Query(None, description="Comma-separated: EL,PL,UPL"),
    status: Optional[str] = Query(None, description="Comma-separated: pending,approved,rejected,emergency,cancelled,cancellation_requested,cancellation_rejected"),
    search: Optional[str] = Query(None, description="Search by name, email, reason, comment"),
    export: bool = Query(False, description="Export as Excel spreadsheet"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service),
):
    roles = current_user.get("roles", [])
    email = current_user.get("email")
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(status_code=403, detail="Admin, PM, or PC access required.")

    from app.features.redmine.sql_service import RedmineSQLService
    sql = RedmineSQLService(leave_service.leave_db.db)
    rm_user = sql.get_user_by_email(email)
    if not rm_user:
        raise HTTPException(status_code=404, detail="Current user not found in Redmine.")
    member_rm_ids = sql.get_team_member_ids(rm_user["id"])
    member_rm_ids.add(rm_user["id"])
    emp_q = leave_service.leave_db.db.query(EmployeeMaster).filter(
        EmployeeMaster.redmine_user_id.in_(member_rm_ids)
    ).all()
    member_emails = [e.user_email for e in emp_q if e.user_email]
    if not member_emails:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "total_pages": 0}

    result = await leave_service.get_leaves_paginated(
        emails=member_emails,
        from_date=from_date, to_date=to_date,
        leave_type=leave_type, status=status,
        search=search, include_name_search=True,
        sort_by=sort_by, sort_dir=sort_dir,
        page=page, page_size=page_size, export=export,
    )
    if export:
        output = _build_leaves_excel(result["records"])
        audit_logger.info(
            f"Team leave history exported by {current_user.get('username')}",
            extra={
                "correlation_id": request.state.correlation_id,
                "extra_data": {
                    "action": "export_team_leaves",
                    "username": current_user.get("username"),
                    "status": "success",
                    "client_ip": _get_client_ip(request),
                    "exported_rows": len(result["records"]),
                },
            },
        )
        filename = f"team_leaves_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    return result


@router.get("/holidays")
async def get_holidays(
    request: Request,
    search: str = Query(None, description="Search by holiday name"),
    holiday_type: str = Query(None, description="Filter: GAZETTED, RESTRICTED"),
    is_national: bool = Query(None, description="Filter by national flag"),
    export: bool = Query(False, description="Export as Excel spreadsheet"),
    page: int = Query(1, description="Page number"),
    page_size: int = Query(50, description="Results per page"),
    sort_by: str = Query("holiday_date", description="Sort column: holiday_date, holiday_name, holiday_type, country_code"),
    sort_dir: str = Query("asc", description="Sort direction: asc, desc"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    page = max(1, page)
    page_size = max(1, min(100, page_size))
    search_term = (search or "").strip()[:100]

    query = db.query(HolidayModel)
    if holiday_type:
        query = query.filter(HolidayModel.holiday_type == holiday_type.upper())
    if is_national is not None:
        query = query.filter(HolidayModel.is_national == is_national)
    if search_term:
        query = query.filter(HolidayModel.holiday_name.ilike(f"%{search_term}%"))

    SORT_COLUMNS = {"holiday_date", "holiday_name", "holiday_type", "country_code"}
    sort_by = sort_by if sort_by in SORT_COLUMNS else "holiday_date"
    sort_dir = sort_dir if sort_dir in ("asc", "desc") else "asc"

    query = query.order_by(
        getattr(HolidayModel, sort_by).desc() if sort_dir == "desc"
        else getattr(HolidayModel, sort_by).asc()
    )
    total = query.count()

    if export:
        holidays = query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Holidays"
        ws.append(["Country Code", "Region", "Date", "Holiday Name", "Type", "National"])

        for h in holidays:
            ws.append([
                h.country_code,
                h.region or "",
                h.holiday_date.isoformat() if h.holiday_date else "",
                h.holiday_name,
                h.holiday_type,
                "Yes" if h.is_national else "No",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        audit_logger.info(
            f"Holidays exported by {current_user.get('username')}",
            extra={
                "correlation_id": request.state.correlation_id,
                "extra_data": {
                    "action": "export_excel",
                    "username": current_user.get("username"),
                    "status": "success",
                    "client_ip": _get_client_ip(request),
                    "exported_rows": len(holidays),
                },
            },
        )

        filename = f"holidays_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    offset = (page - 1) * page_size
    holidays = query.offset(offset).limit(page_size).all()

    return {
        "holidays": [h.to_dict() for h in holidays],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
    }

@router.delete("/holidays/{holiday_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_holiday(
    holiday_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_admin),
):
    holiday = db.query(HolidayModel).filter(HolidayModel.id == holiday_id).first()
    if not holiday:
        raise HTTPException(status_code=404, detail="Holiday not found.")
    db.delete(holiday)
    db.commit()

from fastapi import UploadFile, File

@router.post("/holidays/upload")
async def upload_holidays(
    request: Request,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_admin),
    file: UploadFile = File(...),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    """Upload Excel file to sync holidays. (Admins only)"""
    correlation_id = request.state.correlation_id
    try:
        MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")

        # Validate file size (Limit to 5MB)
        if file.size and file.size > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail="File size exceeds the 5MB limit.")

        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail="File content exceeds the 5MB limit.")

        result = await leave_service.upload_holidays_from_excel(contents)
        return result
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error uploading holidays: {e}")
        raise HTTPException(status_code=500, detail="Error uploading holidays")
    finally:
        await file.close()


@router.get("/holidays/template")
@limiter.limit("5/minute")
async def download_holiday_template(
    request: Request,
    year: int = Query(..., ge=2024, le=2030, description="Year (2024-2030)"),
    country: str = Query("IN", min_length=2, max_length=2, pattern=r"^[A-Z]{2}$", description="Country code, e.g. IN"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_admin),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service),
):
    result = await leave_service.generate_holiday_template(year, country.upper())
    filename = f"holidays_{year}_{country.upper()}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        result["excel"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/validate")
async def validate_leave_dates(
    payload: dict,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    return await leave_service.validate_leave_dates(
        datetime.fromisoformat(payload["start_date"]),
        datetime.fromisoformat(payload["end_date"]),
        current_user.get("sub"),
        [datetime.fromisoformat(d) for d in payload["leave_dates"]] if "leave_dates" in payload else None,
    )

@router.post("/apply")
@limiter.limit("5/minute")
async def apply_leave(
    request: Request,
    leave_data: LeaveApplyRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    user_id = current_user.get("sub")
    email = current_user.get("email")
    result = await leave_service.apply_for_leave(user_id, email, leave_data)
    leave_id = result["leave_id"]
    bal = result["balance"]

    return {"message": "Leave application submitted successfully", "leave_id": leave_id, "leave_ids": result["leave_ids"], "created": result["created"], "balance": bal}


@router.post("/emergency")
@limiter.limit("3/minute")
async def emergency_leave(
    request: Request,
    payload: Optional[dict] = None,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    user_id = current_user.get("sub")
    email = current_user.get("email")
    reason = payload.get("reason") if payload else None
    return leave_service.emergency_leave(user_id, email, reason)

@router.get("/pending")
async def get_pending_leaves(
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    """
    Get all pending leave requests.
    - PM: Gets pending leaves for TRs in their projects.
    - Admin: Gets all pending leaves.
    """
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can view pending leaves."
        )

    pending_list = await leave_service.get_pending_leaves(current_user, from_date, to_date)
    return pending_list


def _build_pending_excel(records: list) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Leaves"
    ws.append(["Date", "Employee", "Designation", "Leave Type", "Request", "Days",
               "Reason", "EL Remaining", "CompOff Remaining"])

    def _fmt(v):
        if v is None:
            return ""
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return str(v)

    for r in records:
        bal = r.get("balance") or {}
        ws.append([
            _fmt(r.get("start_date")),
            r.get("userName") or "",
            r.get("userDesignation") or "",
            r.get("leave_type") or "",
            "Cancellation" if r.get("status") == "cancellation_requested" else "New Leave",
            r.get("requested_days") or 0,
            r.get("reason") or "",
            (bal.get("EL") or {}).get("remaining"),
            (bal.get("CompOff") or {}).get("remaining"),
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/pending-leaves")
async def get_pending_leaves_paginated(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    sort_by: str = Query("created_at", description="created_at, start_date, leave_type, userName"),
    sort_dir: str = Query("asc", description="asc or desc"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    leave_type: Optional[str] = Query(None, description="Comma-separated: EL,PL,UPL"),
    request_type: Optional[str] = Query(None, description="new or cancellation"),
    search: Optional[str] = Query(None, description="Search by name, email, reason, comment"),
    export: bool = Query(False, description="Export as Excel spreadsheet"),
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service),
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(status_code=403, detail="Admin, PM, or PC access required.")

    result = await leave_service.get_pending_leaves_paginated(
        current_user,
        from_date=from_date, to_date=to_date,
        leave_type=leave_type, request_type=request_type,
        search=search,
        sort_by=sort_by, sort_dir=sort_dir,
        page=page, page_size=page_size, export=export,
    )
    if export:
        output = _build_pending_excel(result["records"])
        audit_logger.info(
            f"Pending leaves exported by {current_user.get('username')}",
            extra={
                "correlation_id": request.state.correlation_id,
                "extra_data": {
                    "action": "export_pending_leaves",
                    "username": current_user.get("username"),
                    "status": "success",
                    "client_ip": _get_client_ip(request),
                    "exported_rows": len(result["records"]),
                },
            },
        )
        filename = f"pending_leaves_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    return result


@router.post("/batch/approve")
async def batch_approve_leaves(
    payload: BatchLeaveRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only Admin, PM, or PC can approve leaves.")
    return await leave_service.batch_approve_leaves(payload.leave_ids, current_user)


@router.post("/batch/reject")
async def batch_reject_leaves(
    payload: BatchLeaveRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only Admin, PM, or PC can reject leaves.")
    return await leave_service.batch_reject_leaves(payload.leave_ids, current_user)


@router.post("/batch/cancel")
async def batch_cancel_leaves(
    payload: BatchLeaveRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    return await leave_service.batch_cancel_leaves(payload.leave_ids, current_user)


@router.post("/batch/approve-cancel")
async def batch_approve_cancel_leaves(
    payload: BatchLeaveRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only Admin, PM, or PC can approve cancellation requests.")
    return await leave_service.batch_approve_cancel_leaves(payload.leave_ids, current_user)


@router.post("/batch/reject-cancel")
async def batch_reject_cancel_leaves(
    payload: BatchCancelLeaveRejectRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only Admin, PM, or PC can reject cancellation requests.")
    return await leave_service.batch_reject_cancel_leaves(payload.leave_ids, current_user, payload.remark)


@router.post("/{leave_id}/approve")
async def approve_leave(
    leave_id: str,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    """
    Approve a leave application.
    - PM: Can approve leaves for TRs in their projects.
    - Admin: Full access.
    """
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can approve leaves."
        )

    success = await leave_service.approve_leave(leave_id, current_user)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Leave application not found or unauthorized."
        )

    return {"message": "Leave application approved successfully"}


@router.post("/{leave_id}/reject")
async def reject_leave(
    leave_id: str,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    """
    Reject a leave application.
    - PM: Can reject leaves for TRs in their projects.
    - Admin: Full access.
    """
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can reject leaves."
        )

    success = await leave_service.reject_leave(leave_id, current_user)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Leave application not found or unauthorized."
        )

    return {"message": "Leave application rejected successfully"}


@router.post("/{leave_id}/request-cancel")
async def request_cancel_leave(
    leave_id: str,
    payload: CancelLeaveRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    return await leave_service.request_cancel_leave(leave_id, current_user, payload.remark)


@router.post("/{leave_id}/approve-cancel")
async def approve_cancel_leave(
    leave_id: str,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can approve cancellation requests."
        )
    return await leave_service.approve_cancel_leave(leave_id, current_user)


@router.post("/{leave_id}/reject-cancel")
async def reject_cancel_leave(
    leave_id: str,
    payload: CancelLeaveRejectRequest,
    current_user: dict = Depends(get_current_user),
    leave_service: LeaveBusinessService = Depends(get_leave_business_service)
):
    roles = current_user.get("roles", [])
    if "Admin" not in roles and "Project Manager" not in roles and "Project Coordinator" not in roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admin, PM, or PC can reject cancellation requests."
        )
    return await leave_service.reject_cancel_leave(leave_id, current_user, payload.remark)


@router.get("/{leave_id}", response_model=LeaveHistoryItem)
async def get_leave_by_id(
    leave_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    try:
        from uuid import UUID
        UUID(leave_id)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=404, detail="Leave not found")

    leave = db.query(Leave).filter(Leave.id == leave_id).first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")

    email = current_user.get("email")
    roles = current_user.get("roles", [])

    if "Admin" in roles:
        pass
    elif leave.user_email == email:
        pass
    elif "Project Manager" in roles or "Project Coordinator" in roles:
        from app.features.redmine.sql_service import RedmineSQLService
        sql = RedmineSQLService(db)
        current_rm = sql.get_user_by_email(email)
        target_rm = sql.get_user_by_email(leave.user_email)
        if not current_rm or not target_rm:
            raise HTTPException(status_code=404, detail="Redmine user not found")
        current_projects = sql.get_projects_for_user(current_rm["id"])
        target_projects = sql.get_projects_for_user(target_rm["id"])
        current_pids = {p.id for p in current_projects}
        target_pids = {p.id for p in target_projects}
        if not current_pids.intersection(target_pids):
            raise HTTPException(status_code=403, detail="Not authorized to view this leave")
    else:
        raise HTTPException(status_code=403, detail="Not authorized to view this leave")

    emp = db.query(EmployeeMaster).filter(EmployeeMaster.user_email == leave.user_email).first()
    data = leave.to_dict()
    lt = db.query(LeaveTypeConfig).filter(LeaveTypeConfig.code == leave.leave_type).first()
    if not lt and leave.leave_type == "PL":
        lt = db.query(LeaveTypeConfig).filter(LeaveTypeConfig.code == "CO").first()
    data["leave_type_name"] = lt.name if lt else leave.leave_type
    if emp:
        data["userName"] = f"{emp.first_name} {emp.middle_name or ''} {emp.last_name}".strip().replace("  ", " ")
        data["userDesignation"] = emp.designation
    else:
        data["userName"] = leave.user_email
        data["userDesignation"] = None
    return data


@router.delete("/{leave_id}")
@limiter.limit("5/minute")
async def delete_leave(
    request: Request,
    leave_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    leave = db.query(Leave).filter(Leave.id == leave_id).first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")

    email = current_user.get("email")
    roles = current_user.get("roles", [])

    if "Admin" in roles:
        pass
    elif leave.user_email == email:
        pass
    elif "Project Manager" in roles or "Project Coordinator" in roles:
        from app.features.redmine.sql_service import RedmineSQLService
        sql = RedmineSQLService(db)
        current_rm = sql.get_user_by_email(email)
        target_rm = sql.get_user_by_email(leave.user_email)
        if not current_rm or not target_rm:
            raise HTTPException(status_code=404, detail="Redmine user not found")
        current_projects = sql.get_projects_for_user(current_rm["id"])
        target_projects = sql.get_projects_for_user(target_rm["id"])
        current_pids = {p.id for p in current_projects}
        target_pids = {p.id for p in target_projects}
        if not current_pids.intersection(target_pids):
            raise HTTPException(status_code=403, detail="Not authorized to delete this leave")
    else:
        raise HTTPException(status_code=403, detail="Not authorized to delete this leave")

    db.delete(leave)
    db.commit()

    audit_logger.info(
        f"Leave {leave_id} deleted by {current_user.get('username')} (user: {leave.user_email}, type: {leave.leave_type}, status: {leave.approval_status})",
        extra={
            "correlation_id": request.state.correlation_id,
            "extra_data": {
                "action": "delete_leave",
                "username": current_user.get("username"),
                "leave_id": leave_id,
                "leave_user_email": leave.user_email,
                "leave_type": leave.leave_type,
                "leave_status": leave.approval_status,
                "client_ip": _get_client_ip(request),
            },
        },
    )

    return {"message": "Leave deleted", "leave_id": leave_id}
